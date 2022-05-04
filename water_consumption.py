"""
water_consumption.py
Arquivo contém a classe de consumo de água.
"""

import os
import wx
from openpyxl import load_workbook
import csv
import conversor
import database
import file_manager as fm
import data_processing as dp
import database
import global_variables as gv
import graph_drawing
import wx.grid as gridlib
import table

class CreateWaterWindow(wx.Frame):
    """ Cria o frame basico para a inicializacao do app. """

    def __init__(self, parent):
        style = wx.DEFAULT_FRAME_STYLE & (~wx.MAXIMIZE_BOX) & (~wx.RESIZE_BORDER)
        super().__init__(parent, style=style)

        self.defaultRowSize = 24
        self.WINDOW_NAME = 'Consumo de Água'

        self.table = table.Table(self, self.defaultRowSize, 3)
        self.status_bar = self.CreateStatusBar()
        self.menu = wx.MenuBar()
        self.toolbar = wx.ToolBar(self)
        self.quick_toolbar = wx.ToolBar(self)

        # Inicialização das variaveis.
        self.isSaved = True
        self.isErrors = False
        self.CellErrorsList = []

        self.dataWindow = None      # Janela para adicionar e removar datas.
        self.conversorWindow = None # Janela do conversor de unidades
        self.databaseWindow = None  # Janela do banco de dados.
        self.calendarWindow = None  # "Janela" de exibição de gráficos
        self.lastDate = ''          # Ultima data adicionada na tabela.

        self.OnInit(self.menu)

        # E uma lista de dicionarios que contem as informacoes de cada linha.
        # Exemplo: [{'date': '09/03/2021', 'time': '12:30', 'value': '12.9'}, ...]
        self.data = []
        self.LoadFile()

        # Bind do evento quando o usuário clica no X da janela para sair.
        self.Bind(wx.EVT_CLOSE, self.close_app)
        self.CenterOnScreen()

    def OnInit(self, menu):
        """ Inicializa a UI. """

        # Cria os menus.
        self.setup_arquivo_menu(menu)
        self.setup_tabela_menu(menu)
        self.setup_tools_menu(menu)
        self.SetMenuBar(menu)

        # Cria a toolbar.
        self.setup_toolbar(self.toolbar)

        # Cria a toolbar secundária
        self.setup_quickToolbar(self.quick_toolbar)

        # Sizers para as toolbars e a tabela definida em table.py
        self.masterSizer = wx.BoxSizer(wx.VERTICAL)
        hBox = wx.BoxSizer(wx.HORIZONTAL)
        tableBox = wx.BoxSizer(wx.VERTICAL)
        self.rightSizer = wx.BoxSizer(wx.VERTICAL)
        self.initTextBox()
        self.initGraphBox()

        self.masterSizer.Add(self.toolbar, flag=wx.EXPAND)
        self.masterSizer.Add(self.quick_toolbar, flag=wx.EXPAND)
        tableBox.Add(self.table.panel)

        hBox.Add(tableBox)
        hBox.Add(self.rightSizer, flag=wx.LEFT | wx.EXPAND, border=1)
        self.masterSizer.Add(hBox)

        self.SetSizerAndFit(self.masterSizer)

        # self.SetSize((782, 640))
        self.SetMinSize((782, 640))
        self.SetMaxSize((782, 640))
        self.SetTitle(f'{self.WINDOW_NAME}')

    def setup_arquivo_menu(self, menu):
        """ Inicializa o menu 'Arquivo' do programa. """

        # Inicializa os itens do Menu 'Arquivo'.
        file_menu = wx.Menu()
        new_menu_item = file_menu.Append(wx.ID_NEW, '&Novo\tCtrl+N', 'Criar um novo gráfico')
        openfile_menu_item = file_menu.Append(wx.ID_OPEN, 'Abrir\tCtrl+O', 'Abrir um arquivo')
        save_menu_item = file_menu.Append(wx.ID_SAVE, '&Salvar\tCtrl+S', 'Salvar o gráfico')
        saveAs_menu_item = file_menu.Append(wx.ID_SAVEAS, 'Salvar como...', 'Salvar o gráfico em um arquivo diferente')

        # Criando o sub-menu Importar...
        importar = wx.Menu()
        import_excel = importar.Append(wx.ID_ANY, 'Arquivo excel', 'Importar um arquivo excel. Limpa toda a tabela')
        import_csv = importar.Append(wx.ID_ANY, 'Arquivo csv', 'Importar um arquivo csv. Limpa toda a tabela')
        import_txt = importar.Append(wx.ID_ANY, 'Arquivo txt', 'Importar um arquivo txt. Limpa toda a tabela')
        file_menu.AppendSubMenu(importar, 'Importar...')

        close_menu_item = file_menu.Append(wx.ID_CLOSE, '&Fechar\tCtrl+F', 'Fecha o arquivo aberto')
        draw_menu_item = file_menu.Append(wx.ID_ANY, '&Desenhar gráfico\tCtrl+D', 'Desenhar o gráfico na tela')
        file_menu.AppendSeparator()
        home_menu_item = file_menu.Append(wx.ID_HOME, 'Voltar pra Home\tCtrl+H', 'Voltar para a tela de boas vindas.')
        quit_menu_item = file_menu.Append(wx.ID_EXIT, 'Sair\tCtrl+Q', 'Sair do programa')

        # Binding para os itens do menu.
        self.Bind(wx.EVT_MENU, lambda event: self.OnQuit(event, is_exit=True), quit_menu_item)
        self.Bind(wx.EVT_MENU, self.OnSave, save_menu_item)
        self.Bind(wx.EVT_MENU, self.OnSaveAs, saveAs_menu_item)
        self.Bind(wx.EVT_MENU, self.LoadExcelFile, import_excel)
        self.Bind(wx.EVT_MENU, self.LoadCsvFile, import_csv)
        self.Bind(wx.EVT_MENU, self.LoadTxtFile, import_txt)
        self.Bind(wx.EVT_MENU, self.OnOpenFile, openfile_menu_item)
        self.Bind(wx.EVT_MENU, self.OnCloseFile, close_menu_item)
        self.Bind(wx.EVT_MENU, self.OnNewFile, new_menu_item)
        self.Bind(wx.EVT_MENU, self.OnDraw, draw_menu_item)
        self.Bind(wx.EVT_MENU, lambda event: self.OnQuit(event, is_exit=False), home_menu_item)

        self.Bind(gridlib.EVT_GRID_CELL_CHANGED, self.OnCellChange)

        # Adicionar o Menu a barra superior.
        menu.Append(file_menu, '&Arquivo')

    def setup_tabela_menu(self, menu):
        """ Inicializa o menu 'Tabela' """

        tabela_menu = wx.Menu()

        clear_table_item = tabela_menu.Append(wx.ID_ANY, 'Limpar tabela', 'Limpa toda a tabela')
        fill_data_item = tabela_menu.Append(wx.ID_ANY, 'Adicionar data', 'Preencher fileiras com Datas e Horários')
        remove_data_item = tabela_menu.Append(wx.ID_ANY, 'Remover data', 'Remove a última data da tabela.')

        self.Bind(wx.EVT_MENU, self.clearTable, clear_table_item)
        self.Bind(wx.EVT_MENU, self.OnAddDataWindow, fill_data_item)
        self.Bind(wx.EVT_MENU, self.OnDeleteDataWindow, remove_data_item)

        menu.Append(tabela_menu, 'Tabela')


    def setup_tools_menu(self, menu):
        ''' Inicializa o menu de 'Ferramentas'. '''

        file_menu = wx.Menu()

        conversor_item_menu = file_menu.Append(-1, '&Conversor', 'Abrir o conversor de unidades')
        database_item_menu = file_menu.Append(-1, '&Banco de dados', 'Abrir o banco de dados')

        self.Bind(wx.EVT_MENU, self.OnConversor, conversor_item_menu)
        self.Bind(wx.EVT_MENU, self.OnDatabase, database_item_menu)

        menu.Append(file_menu, 'Ferramentas')

        self.SetMenuBar(menu)


    def setup_toolbar(self, toolbar):
        """ Inicializa a toolbar. """

        new_tool = toolbar.AddTool(wx.ID_NEW, 'Novo', wx.Bitmap('images/new.png'), 'Novo arquivo')
        open_tool = toolbar.AddTool(wx.ID_OPEN, 'Abrir', wx.Bitmap('images/open.png'), 'Abrir arquivo')
        save_tool = toolbar.AddTool(wx.ID_SAVE, 'Salvar', wx.Bitmap('images/save.png'), 'Salvar arquivo')
        draw_tool = toolbar.AddTool(wx.ID_ANY, 'Desenhar Gráfico', wx.Bitmap('images/graph.png'), 'Desenhar gráfico')
        conversor_tool = toolbar.AddTool(wx.ID_ANY, 'Conversor de unidades', wx.Bitmap('images/calculator.png'), 'Conversor de unidades')
        database_tool = toolbar.AddTool(wx.ID_ANY, 'Banco de Dados', wx.Bitmap('images/database.png'), 'Banco de dados')
        home_tool = toolbar.AddTool(wx.ID_HOME, 'Voltar pra Home', wx.Bitmap('images/home.png'), 'Voltar para Home')

        # Funcoes de Binding. Estamos usando as mesmas para os itens do menu.
        self.Bind(wx.EVT_TOOL, self.OnNewFile, new_tool)
        self.Bind(wx.EVT_TOOL, self.OnOpenFile, open_tool)
        self.Bind(wx.EVT_TOOL, self.OnSave, save_tool)
        self.Bind(wx.EVT_TOOL, self.OnDraw, draw_tool)
        self.Bind(wx.EVT_TOOL, self.OnConversor, conversor_tool)
        self.Bind(wx.EVT_TOOL, self.OnDatabase, database_tool)
        self.Bind(wx.EVT_TOOL, lambda event: self.OnQuit(event, is_exit=False), home_tool)

        toolbar.Realize()

    def setup_quickToolbar(self, toolbar):
        ''' Inicializa a toolbar secundaria. '''

        add_table = toolbar.AddTool(wx.ID_ANY, 'Inserir dados', wx.Bitmap('images/add.png'), 'Adicionar data')
        delete_table = toolbar.AddTool(wx.ID_ANY, 'Deletar dados', wx.Bitmap('images/remove.png'), 'Remover data')

        self.Bind(wx.EVT_TOOL, self.OnAddDataWindow, add_table)
        self.Bind(wx.EVT_TOOL, self.OnDeleteDataWindow, delete_table)

        toolbar.Realize()

    def initTextBox(self):
        ''' Inicializa a self.textBox com um texto padrão. '''

        self.textBox = wx.BoxSizer(wx.VERTICAL)

        msg = "Entre com dados de consumo e clique em 'Desenhar Gráfico' para a exibição.\n\n"
        msg += "Não se esqueça de clicar novamente em 'Desenhar Gráfico' quando modificar alguma informação para atualizar este painel."
        text = wx.StaticText(self, -1, msg, size=(250, 400))
        text.SetFont(wx.Font(13, wx.FONTFAMILY_ROMAN, wx.DEFAULT, wx.DEFAULT))
        self.textBox.Add(text, flag=wx.TOP, border=25)

        self.rightSizer.Add(self.textBox, flag=wx.ALL, border=25)

    def initGraphBox(self):
        ''' Inicializa a GraphBox. '''

        self.graphBox = wx.BoxSizer(wx.VERTICAL)
        self.calendarWindow = graph_drawing.GraphCalendar(self, None)
        self.graphBox.Add(self.calendarWindow)
        self.graphBox.ShowItems(False)
        self.rightSizer.Add(self.graphBox)

    def OnQuit(self, event, is_exit):
        """ Chamada quando o usuário tentar sair do app.
            ``@is_exit`` diz se o usuário quer sair ou voltar pra tela de boas vindas.
            True para sair, False para voltar para a janela de boas vindas. """

        if not self.isSaved:
            if gv.opened_file:
                msg = f'Desejar salvar o arquivo "{gv.filename}" antes de sair?'
            else:
                msg = 'Desejar salvar antes de sair?'

            message = wx.MessageDialog(self, msg, 'Arquivo não salvo.', wx.YES_NO | wx.CANCEL | wx.ICON_QUESTION)
            answer = message.ShowModal()
            if answer == wx.ID_CANCEL:
                return

            if answer == wx.ID_YES:
                if gv.opened_file:
                    self.SaveFile()
                    if is_exit:
                        self.CloseFile()
                else:
                    if self.OnSave(event):
                        if is_exit:
                            self.CloseFile()
                    else:
                        return

        self.Destroy()
        if is_exit:
            gv.welcome_screen.destroy_()
            self.Layout()
        else:
            gv.welcome_screen.show_()

    def OnSave(self, event):
        """ Chamada quando o usuário tentar salvar o arquivo. Retorna True em caso de sucesso. """

        if self.isThereErrorsWithTable():
            return

        if gv.opened_file:
            self.SaveFile()
            return True

        # Se o usuário estiver criando um arquivo do zero...
        else:
            dialog = wx.FileDialog(self, f"Escolha um nome para o arquivo {gv.file_suffix}", gv.file_dir, '', f'*{gv.file_suffix}*', wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)

            if dialog.ShowModal() == wx.ID_OK:
                fm.saveAndUpdateFileVariables(dialog)

                # Cria o arquivo, apenas.
                fm.createEmptyFile()

                gv.opened_file = open(gv.file_path, 'r')
                self.isSaved = True
                self.SetTitle(f'{gv.filename} - {self.WINDOW_NAME}')

                dialog.Destroy()
                self.table.transferTableToList(self.data)
                self.writeWaterComsumptionToFile()

                return True

            return False

    def OnOpenFile(self, event):
        """ Chamada quando o usuário tentar abrir um arquivo. """

        if not self.isSaved:
            if gv.opened_file:
                msg = f'Desejar salvar o arquivo "{gv.filename}" antes de abrir um novo?'
            else:
                msg = 'Desejar salvar antes de abrir um novo?'

            message = wx.MessageDialog(self, msg, 'Arquivo não salvo.', wx.YES_NO | wx.CANCEL | wx.ICON_QUESTION)
            answer = message.ShowModal()
            if answer == wx.ID_CANCEL:
                return

            if gv.opened_file:
                if answer == wx.ID_YES:
                    self.SaveFile()
                    self.CloseFile()
                else:
                    self.CloseFile()


        dialog = wx.FileDialog(self, f"Escolha um arquivo {gv.file_suffix}", gv.file_dir, '', f'*{gv.file_suffix}*', wx.FD_OPEN)

        # Se a ID do Modal for OK, significa que o usuário escolheu um arquivo.
        if dialog.ShowModal() == wx.ID_OK:
            self.CloseFile()
            fm.openAndUpdateFileVariables(dialog)
            isFileOK = fm.isFileIntegrityOK()
            self.data = fm.copyWaterFileDataToList()

            # Se a lista estiver vazia (não há dados de consumo de água), tudo OK, só não há dados.
            if isFileOK and not self.data:
                dialog.Destroy()
                self.table.paintAllBlank()
                self.isErrors = False
                self.isSaved = True
                self.SetTitle(f'{gv.filename} - {self.WINDOW_NAME}')
                return

            if isFileOK and fm.analizeWaterData(self.data):
                self.table.rearrangeRows(len(self.data))
                self.table.transferListToTable(self.data)
                self.table.paintAllBlank()
                self.isErrors = False
                self.isSaved = True
                self.SetTitle(f'{gv.filename} - {self.WINDOW_NAME}')

                self.OnDraw(None)

            else:
                fm.clearFileVariables()
                dlg = wx.MessageDialog(self, 'O arquivo possui erros.\nEle pode ter sido modificado inadvertidamente ou corrompido.',
                'Erro com o arquivo', wx.OK | wx.ICON_ERROR)
                dlg.ShowModal()

        dialog.Destroy()

    def OnSaveAs(self, event):
        """ Funcao e chamada quando o usuário clica em "Salvar como..." """

        dialog = wx.FileDialog(self, "Escolha um nome para o arquivo", gv.file_dir, '', f'*{gv.file_suffix}*', wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)

        if dialog.ShowModal() == wx.ID_OK:
            fm.saveAndUpdateFileVariables(dialog)

            fm.createEmptyFile()

            self.isSaved = True
            self.SetTitle(f'{gv.filename} - {self.WINDOW_NAME}')

            dialog.Destroy()
            self.table.transferTableToList(self.data)
            self.writeWaterComsumptionToFile()


    def LoadExcelFile(self, event):
        """ Carrega um arquivo excel com a extensao .xlsx. """

        if gv.opened_file:
            dial = wx.MessageDialog(self, 'Todos os dados da tabela serão perdidos. Deseja continuar?', 'Tem certeza?', wx.YES_NO | wx.ICON_QUESTION)
            result = dial.ShowModal()

            if result != wx.ID_YES:
                return

        dialog = wx.FileDialog(self, "Escolha um arquivo .xlsx", '', '', '*.xlsx', wx.FD_OPEN)

        # Se a ID do Modal for OK, significa que o usuário escolheu um arquivo.
        if dialog.ShowModal() == wx.ID_OK:
            filename = dialog.GetFilename()
            file_dir = dialog.GetDirectory()
            file_path = os.path.join(file_dir, filename)
        else:
            return

        # Nos não éamos fazer modicacoes no arquivo excel, so precisamos saber o path para pegarmos os dados.
        workbook = load_workbook(file_path, data_only=True)

        sheet = workbook.active # Aqui a gente diz que vai trabalhar com a primeira "folha" do excel.
        dic = {}

        self.data.clear()
        self.table.clearTable()
        try:
            for i in range(2, sheet.max_row + 1): # Os dados comecam do index row=2.
                for j in range(1, 4):
                    if j == 1:
                        dic['date'] = dp.excelDateToOurDate(str(sheet.cell(i, j).value))
                    if j == 2:
                        dic['time'] = dp.excelTimeToOurTime(str(sheet.cell(i, j).value))
                    if j == 3:
                        dic['value'] = str(sheet.cell(i, j).value)

                self.data.append(dic)
                dic = {}

            if not fm.analizeWaterData(self.data):
                raise Exception('File error')

            self.table.rearrangeRows(len(self.data))
            self.table.transferListToTable(self.data)
            self.table.paintAllBlank()
            self.isErrors = False
            self.isSaved = False
            if gv.opened_file:
                self.SetTitle(f'* {gv.filename} - {self.WINDOW_NAME}')
            else:
                self.SetTitle(f'(Não Salvo) - {self.WINDOW_NAME}')

            self.OnDraw(None)

        except:
            self.data.clear()
            self.table.clearTable()

            dial = wx.MessageDialog(self, 'Erro no processamento do arquivo. Ele pode não estar no formato ou identação correta.', 'Erro encontrado.', wx.ICON_ERROR)
            dial.ShowModal()

    def LoadCsvFile(self, event):
        ''' Carrega um arquivo .csv para a tabela '''

        if gv.opened_file:
            dial = wx.MessageDialog(self, 'Todos os dados da tabela serão perdidos. Deseja continuar?', 'Tem certeza?', wx.YES_NO | wx.ICON_QUESTION)
            result = dial.ShowModal()

            if result != wx.ID_YES:
                return

        dialog = wx.FileDialog(self, "Escolha um arquivo .csv", '', '', '*.csv', wx.FD_OPEN)

        # Se a ID do Modal for OK, significa que o usuário escolheu um arquivo.
        if dialog.ShowModal() == wx.ID_OK:
            filename = dialog.GetFilename()
            file_dir = dialog.GetDirectory()
            file_path = os.path.join(file_dir, filename)
        else:
            return

        dic = {}
        self.data.clear()
        self.table.clearTable()

        try:
            with open(file_path, 'r') as f:
                csvRead = csv.reader(f, delimiter=',')
                next(csvRead, None)

                for row in csvRead:
                    dic['date'] = row[0]
                    dic['time'] = row[1]
                    dic['value'] = row[2]
                    self.data.append(dic)
                    dic = {}

                if not fm.analizeWaterData(self.data):
                    raise Exception('File error')

                self.table.rearrangeRows(len(self.data))
                self.table.transferListToTable(self.data)
                self.table.paintAllBlank()
                self.isErrors = False
                self.isSaved = False
                self.SetTitle(f'(Não Salvo) - {self.WINDOW_NAME}')

                self.OnDraw(None)
        except:
            self.data.clear()
            self.table.clearTable()

            dial = wx.MessageDialog(self, 'Erro no processamento do arquivo. Ele pode não estar no formato ou identação correta.', 'Erro encontrado', wx.ICON_ERROR)
            dial.ShowModal()


    def LoadTxtFile(self, event):
        ''' Importa um arquivo .txt. '''

        if gv.opened_file:
            dial = wx.MessageDialog(self, 'Todos os dados da tabela serão perdidos. Deseja continuar?', 'Tem certeza?', wx.YES_NO | wx.ICON_QUESTION)
            result = dial.ShowModal()

            if result != wx.ID_YES:
                return

        dialog = wx.FileDialog(self, "Escolha um arquivo .txt", '', '', '*.txt', wx.FD_OPEN)

        # Se a ID do Modal for OK, significa que o usuário escolheu um arquivo.
        if dialog.ShowModal() == wx.ID_OK:
            filename = dialog.GetFilename()
            file_dir = dialog.GetDirectory()
            file_path = os.path.join(file_dir, filename)
        else:
            return

        dic = {}
        self.data.clear()
        self.table.clearTable()

        try:
            txtFile = open(file_path, 'r')
            lines = txtFile.readlines()

            for line in lines:
                words = line.split()

                dic['date'] = words[0]
                dic['time'] = words[1]
                dic['value'] = words[2]

                self.data.append(dic)
                dic = {}

            if not fm.analizeWaterData(self.data):
                raise Exception('File error')

            self.table.rearrangeRows(len(self.data))
            self.table.transferListToTable(self.data)
            self.table.paintAllBlank()
            self.isErrors = False
            self.isSaved = False
            self.SetTitle(f'(Não Salvo) - {self.WINDOW_NAME}')


            txtFile.close()
            self.OnDraw(None)

        except:
            self.data.clear()
            self.table.clearTable()

            dial = wx.MessageDialog(self, 'Erro no processamento do arquivo. Ele pode não estar no formato ou identação correta.', 'Erro encontrado', wx.ICON_ERROR)
            dial.ShowModal()


    def OnNewFile(self, event):
        """ Chamada quando o usuário tentar criar um novo arquivo. """

        if not self.isSaved:
            if gv.opened_file:
                msg = f'Desejar salvar o arquivo "{gv.filename}" antes de criar um novo?'
                message = wx.MessageDialog(self, msg, 'Arquivo não salvo.', wx.YES_NO | wx.CANCEL | wx.ICON_QUESTION)
                answer = message.ShowModal()
                if answer == wx.ID_CANCEL:
                    return

                if answer == wx.ID_YES:
                    self.SaveFile()
                    self.CloseFile()
                else:
                    self.CloseFile()

            else:
                msg = 'Desejar salvar o arquivo antes de criar um novo?'
                message = wx.MessageDialog(self, msg, 'Arquivo não salvo.', wx.YES_NO | wx.CANCEL | wx.ICON_QUESTION)
                answer = message.ShowModal()
                if answer == wx.ID_CANCEL:
                    return

                elif answer == wx.ID_NO:
                    self.CloseFile()
                    self.table.paintAllBlank()

                else:
                    self.OnSave(event)
                    self.table.paintAllBlank()

                self.clearGraphBox()

        else:
            if gv.opened_file:
                self.CloseFile()

    def OnCloseFile(self, event):
        """ Chamada quando o usuário tenta fechar o arquivo. """

        if self.isSaved:
            self.CloseFile()
        else:
            msg = f'Desejar salvar antes de fechar?'
            if gv.opened_file:
                msg = f'Desejar salvar o arquivo "{gv.filename}" antes de fechar?'

            message = wx.MessageDialog(self, msg, 'Arquivo não salvo.', wx.YES_NO | wx.CANCEL | wx.ICON_QUESTION)
            answer = message.ShowModal()
            if answer == wx.ID_CANCEL:
                return

            if answer == wx.ID_YES:
                self.SaveFile()
                self.CloseFile()
            else:
                self.CloseFile()

            self.table.clearTable()

        event.Skip()

    def OnDraw(self, event):
        """ Chamada quando o usuário tenta desenhar. """

        if self.isThereErrorsWithTable():
            return

        if not self.data:
            dlg = wx.MessageDialog(self, 'Por favor, verifique a tabela por células vazias.', 'Erro encontrado', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            return

        self.table.transferTableToList(self.data)
        daysList = dp.getTableReadyData(self.data)

        self.calendarWindow.data = daysList
        self.calendarWindow.OrganizeData()
        self.textBox.ShowItems(False)
        self.graphBox.ShowItems(True)
        self.rightSizer.Layout()


    def SaveOverallGraph(self, path):
        ''' Ao invés de exibir na tela, salva em ``dir`` o gráfico de consumo geral e destroi a janela. '''

        self.table.transferTableToList(self.data)
        daysList = dp.getTableReadyData(self.data)
        self.calendarWindow.data = daysList

        self.calendarWindow.isSaveToDisk = True
        self.calendarWindow.path = path
        self.calendarWindow.OnSummary(None)

    def SaveFile(self):
        """ Responsável por salvar o arquivo. Supõe que já existe um arquivo aberto. """

        if self.isThereErrorsWithTable():
            return

        if gv.opened_file and not self.isSaved:
            self.table.transferTableToList(self.data)
            self.writeWaterComsumptionToFile()
            self.isSaved = True
            self.SetTitle(f'{gv.filename} - {self.WINDOW_NAME}')

        elif not gv.opened_file:
            self.OnSaveAs(None)


    def LoadFile(self):
        """ Carrega os dados do arquivo para as tabelas. E chamado para inicializacao apenas. """

        if gv.opened_file:
            if '#water_consumption_start' in gv.fileStartIndices.keys():
                self.data = fm.copyWaterFileDataToList()
                self.table.rearrangeRows(len(self.data))
                self.table.transferListToTable(self.data)
                self.isErrors = False
                self.isSaved = True
                self.SetTitle(f'{gv.filename} - {self.WINDOW_NAME}')
                self.OnDraw(None)
            else:
                self.isSaved = True
                self.SetTitle(f'{gv.filename} - {self.WINDOW_NAME}')

    def CloseFile(self):
        """ Limpa a tabela e fecha o arquivo aberto. Pode ser chamada mesmo não éabendo
        se ha um arquivo aberto. """

        if gv.opened_file:
            self.SetTitle(f'{self.WINDOW_NAME}')
            fm.clearFileVariables()
            self.isSaved = True
            self.table.clearTable()
            self.table.deleteRows(self.table.rowLength - self.defaultRowSize)

        # Se não éiver arquivo aberto, apenas "limpa" a janela.
        else:
            self.SetTitle(f'{self.WINDOW_NAME}')
            self.table.clearTable()
            self.table.deleteRows(self.table.rowLength - self.defaultRowSize)
            self.isSaved = True

        self.table.paintAllBlank()
        self.CellErrorsList.clear()
        self.table.lastFilledRow = 0

        self.graphBox.ShowItems(False)
        self.textBox.ShowItems(True)


    def updateTitleName(self):
        """ Atualiza o nome da janela. Normalmente chamada quando houver mudanca nas células. """

        if gv.opened_file:
            if self.isSaved:
                msg = f'{gv.filename} - {self.WINDOW_NAME}'
            else:
                msg = f'* {gv.filename} - {self.WINDOW_NAME}'
        else:
            msg = f'(Não Salvo) - {self.WINDOW_NAME}'

        self.SetTitle(msg)


    def OnCellChange(self, event):
        """ Chamada quando há alguma modificação em uma célula. """

        self.isSaved = False
        self.updateTitleName()

        row = event.GetRow()
        col = event.GetCol()
        t = (row, col)

        if self.table.checkConsumoValue(row):
            self.isErrors = self.ManageErrorsWithCells(True, row)
        else:
            self.isErrors = self.ManageErrorsWithCells(False, row)


    def writeWaterComsumptionToFile(self):
        ''' Salva toda a tabela de consumo de água no arquivo. '''

        before = []

        # Sempre vamos colocar nossos dados no inicio do arquivo. Se dados de água ja existirem, vamos apaga-lo totalmente e reescrever os de agora.
        # Se não,évamos colocar esses dados no inicio e todo o resto depois.
        if '#water_consumption_start' in gv.fileStartIndices.keys():
            before = gv.fileLines[ : gv.fileStartIndices['#water_consumption_start'] - 1]
            after = gv.fileLines[gv.fileStartIndices['#water_consumption_end'] + 1 : ] # O indice de fim aponta exatamente para o indicador.
        else:
            after = gv.fileLines

        outText = before
        outText += ['#water_consumption_start\n']
        for dic in self.data:
            outText.append(f"{dic['date']} {dic['time']} {dic['value']}\n")

        outText += '#water_consumption_end\n'
        outText += after

        # Fechamos o arquivo e o apagamos totalmente.
        gv.opened_file.close()
        gv.opened_file = open(gv.file_path, 'w')

        # Reescrevemos do zero.
        gv.opened_file.write(''.join(outText))
        gv.opened_file.close()

        # Reabrimos em modo de leitura.
        gv.opened_file = open(gv.file_path, 'r')

        fm.getEntrysDictAndFile()


    def ManageErrorsWithCells(self, isIn, row):
        """ Mantém a lista self.CellErrorsList atualizada.
        Se ``isIn`` for True, row (int) é adicionada na lista, caso False, row é removida. Retorna True se a lista não éstiver vazia. """

        # Se a lista estiver vazia e não é para adicionar novos items, retorna False.
        if not self.CellErrorsList and not isIn:
            return False

        # Precisamos saber se a fileira que vamos adicionar ja esta na lista.
        if isIn:
            if row in self.CellErrorsList:
                return True

            self.CellErrorsList.append(row) # Se chegarmos a esse ponto, cell não éstava na lista
            return True

        else:
            for item in self.CellErrorsList:
                if item == row:
                    self.CellErrorsList.remove(row) # Procura e remove cell.

                    # Se a lista ainda contiver row após a remoção, retorna True
                    if not self.CellErrorsList:
                        return False
                    else:
                        return True

    def checkConsumoValues(self):
        '''Checa todos os dados de consumo. Atualiza a lista self.CellErrorsList. Retorna True se encontrou algum erro. '''

        errorsFound = False

        for row in range(0, self.table.lastFilledRow):
            if self.table.checkConsumoValue(row):
                self.ManageErrorsWithCells(True, row)
                errorsFound = True
            else:
                self.ManageErrorsWithCells(False, row)

        self.table.grid.ForceRefresh()

        return errorsFound

    def isThereErrorsWithTable(self):
        """ Verifica se todos os campos de consumo para onde há dados estão preenchidos corretamente.
        A função colore as células de acordo e exibe mensagem de erro.
        Retorna True se houver erros. """

        thereWasAnError = False

        if self.checkConsumoValues():
            dlg = wx.MessageDialog(self, 'Corriga as células marcadas em vermelho antes de continuar. Consulte o menu de ajuda, se necessário.',
            'Erros encontrados.', wx.ICON_ERROR)
            dlg.ShowModal()

            thereWasAnError = True

        if self.table.lastFilledRow <= 0:
            dlg = wx.MessageDialog(self, 'A tabela não contém nenhuma data. Gere algum dado antes de salvar.', 'Tabela está vazia.',
            wx.ICON_ERROR)
            dlg.ShowModal()

            thereWasAnError = True

        return thereWasAnError

    def close_app(self, event):
        """ Chamada quando o usuário clica no X vermelho para fechar o programa. """

        self.OnQuit(event, is_exit=True)

    def duplicateDate(self, interval, date):
        """ Chama a função generateDate do grid para preencher a tabela com datas. """

        self.table.generateDate(interval, date)
        self.isSaved = False
        self.updateTitleName()

    def duplicateTime(self, interval):
        """ Chama a função generateTime do grid para preencher a tabela com horários. """

        self.table.generateTime(interval)
        self.isSaved = False
        self.updateTitleName()

    def clearTable(self, event):
        ''' Chamada quando o usuário clica para limpar toda a tabela pelo menu Tabela. '''

        self.table.clearTable()
        self.data.clear()
        self.CellErrorsList.clear()
        self.isSaved = False
        self.table.paintAllBlank()
        self.lastDate = ''

        self.table.clearTable()

        if gv.opened_file:
            self.SetTitle(f'* {gv.filename} - {self.WINDOW_NAME}')
        else:
            self.SetTitle(f'(Não Salvo) - {self.WINDOW_NAME}')

    def OnConversor(self, event):
        ''' Chamada quando o usuário clica no botao da calculadora. '''

        if not self.conversorWindow:
            self.conversorWindow = conversor.Conversor(self)
            self.conversorWindow.Show()

    def OnDatabase(self, event):
        ''' Abre a janela do banco de dados. '''

        if not self.databaseWindow:
            self.databaseWindow = database.Database(self)
            self.databaseWindow.Show()

    def OnDatabase(self, event):
        ''' Chamada quando o usuário clicar no ícone de data base na toolbar ou pelos menus. '''

        if not self.databaseWindow:
            self.databaseWindow = database.Database(self)
            self.databaseWindow.Show()

    def OnAddDataWindow(self, event):
        ''' Chamada quando o usuário clica para abrir a janela para adicionar dados a tabela. '''

        if not self.dataWindow:
            self.dataWindow = AddDateTime(self)
            self.dataWindow.Show()

    def OnDeleteDataWindow(self, event):
        ''' Chamada quando o usuário clica para abrir a janela para deletar dados da tabela. '''

        if not self.dataWindow:
            self.dataWindow = DeleteDateTime(self)
            self.dataWindow.Show()


class AddDateTime(wx.Frame):
    ''' Classe responsável pela janela de adicionar Data e Horario. '''

    def __init__(self, parent):
        style = wx.DEFAULT_FRAME_STYLE & (~wx.MAXIMIZE_BOX) & (~wx.RESIZE_BORDER)
        super().__init__(parent, style=style)

        self.parent = parent
        self.SetTitle('Adicionar dados')

        self.setupUI()

        self.Bind(wx.EVT_CLOSE, self.OnCloseWindow)
        self.CenterOnParent()

    def setupUI(self):
        ''' Popula o tableBox com os elementos da janela. '''

        self.masterSizer = wx.BoxSizer(wx.VERTICAL)

        # Adicionando os itens para o tableBox da data.
        dateSizer = wx.BoxSizer(wx.HORIZONTAL)
        dateText = wx.StaticText(self, -1, 'Data ', size=(70, 23))
        self.dateInput = wx.TextCtrl(self, -1, style=wx.TE_PROCESS_ENTER)
        self.dateInput.Bind(wx.EVT_TEXT_ENTER, self.OnAddToGrid)
        self.dateInput.SetMinSize((100, 23))

        dateSizer.Add(dateText, flag=wx.TOP, border=4)
        dateSizer.Add(self.dateInput)
        self.masterSizer.Add(dateSizer, flag=wx.ALL, border=10)

        # Adicionando os itens para o tableBox do intervalo de tempo.
        hourSizer = wx.BoxSizer(wx.HORIZONTAL)
        hourText = wx.StaticText(self, -1, 'Intervalo ', size=(70, 23))
        intervalList = ['1 hora', '30 minutos', '15 minutos', '5 minutos', '1 minuto']
        self.hourSelect = wx.ComboBox(self, -1, choices=intervalList, style=wx.CB_READONLY)
        self.hourSelect.SetMinSize((100, 23))
        self.hourSelect.SetValue('1 hora')

        hourSizer.Add(hourText, flag=wx.TOP, border=4)
        hourSizer.Add(self.hourSelect)
        self.masterSizer.Add(hourSizer, flag=wx.ALL, border=10)

        # Adicionando o botao de adicionar a tabela.
        button = wx.Button(self, -1, 'Adicionar')
        button.Bind(wx.EVT_BUTTON, self.OnAddToGrid)

        self.masterSizer.Add(button, flag=wx.ALIGN_CENTRE | wx.ALL, border=10)
        self.SetSizerAndFit(self.masterSizer)

    def OnAddToGrid(self, event):
        ''' Adiciona os dados de data e horario a tabela. '''

        date = self.dateInput.GetValue()

        if dp.is_date(date):
            self.dateInput.SetBackgroundColour(wx.NullColour)
            self.dateInput.Refresh()
        else:
            self.dateInput.SetBackgroundColour(gv.RED_ERROR)
            self.dateInput.Refresh()
            wx.MessageBox('Data no formato inválido. Por favor, tente novamente.', 'Erro', wx.OK | wx.ICON_ERROR)
            return

        # Verifica se a data e maior que a anterior.
        if not self.parent.lastDate or dp.isDateBigger(self.parent.lastDate, date):
            interval = self.getInterval()
            self.parent.table.rearrangeRows(int(24 * (60 / interval)))
            self.parent.duplicateDate(interval, date)
            self.parent.duplicateTime(interval)
            self.parent.table.updateTimeDateVariables(interval, date)
            self.parent.lastDate = date
            self.parent.isSaved = False
            self.parent.updateTitleName()
            msg = wx.MessageDialog(self, 'Dados inseridos na tabela com sucesso.', 'Sucesso', wx.OK | wx.ICON_INFORMATION)
            msg.ShowModal()

        else:
            self.dateInput.SetBackgroundColour(gv.RED_ERROR)
            self.dateInput.Refresh()
            msg = wx.MessageDialog(self, f'Por favor, escolha uma data posterior a {self.parent.lastDate}.', 'Erro', wx.OK | wx.ICON_ERROR)
            msg.ShowModal()

    def getInterval(self):
        ''' Retorna, em int, o intervalo selecionado na wx.ComboBox do intervalo. '''

        value = self.hourSelect.GetValue()

        if value == '1 hora':
            return 60
        elif value == '30 minutos':
            return 30
        elif value == '15 minutos':
            return 15
        elif value == '5 minutos':
            return 5
        else:
            return 1

    def OnCloseWindow(self, event):
        ''' Funcao chamada quando o usuário clica no botao de fechar no canto superior direito. '''

        self.Destroy()
        self.parent.dataWindow = None


class DeleteDateTime(wx.Frame):
    ''' Classe responsavel pela janela de remoção de dados da tabela. '''

    def __init__(self, parent):
        style = wx.DEFAULT_FRAME_STYLE & (~wx.MAXIMIZE_BOX) & (~wx.RESIZE_BORDER)
        super().__init__(parent, style=style)

        self.parent = parent
        self.SetTitle('Remover dados')
        self.CenterOnParent()

        self.setupUI()

        self.Bind(wx.EVT_CLOSE, self.OnCloseWindow)

    def setupUI(self):
        ''' Inicializa o tableBox principal da UI. '''

        mainSizer = wx.BoxSizer(wx.VERTICAL)

        self.yesBtn = wx.Button(self, -1, 'Sim')
        cancelBtn = wx.Button(self, -1, 'Cancelar')

        if self.parent.lastDate:
            text = f'A última data inserida é de {self.parent.lastDate}.\nVocê deseja removê-la?'
        else:
            text = 'A tabela não contém nenhuma data.'
            self.yesBtn.Enable(False)

        self.message = wx.StaticText(self, -1, text)
        mainSizer.Add(self.message, flag=wx.ALL, border=10)

        btnSizer = wx.BoxSizer(wx.HORIZONTAL)
        btnSizer.Add(self.yesBtn, flag=wx.ALL | wx.CENTRE, border=10)
        btnSizer.Add(cancelBtn, flag=wx.ALL | wx.CENTRE, border=10)

        mainSizer.Add(btnSizer, flag=wx.ALL, border=10)
        self.SetSizerAndFit(mainSizer)

        self.yesBtn.Bind(wx.EVT_BUTTON, self.OnYes)
        cancelBtn.Bind(wx.EVT_BUTTON, self.OnCancel)


    def OnYes(self, event):
        ''' Chamada quando o botao 'Sim' for clicado. '''

        index = self.parent.table.lastFilledRow

        last = self.parent.table.grid.GetCellValue(index -1, 1)
        penultimate = self.parent.table.grid.GetCellValue(index -2, 1)

        diff = dp.getTimesDifference(penultimate, last)
        diff = int(24 * (1 / diff))
        startPos = self.parent.table.lastFilledRow - diff
        previousDate = self.parent.lastDate

        # Deletando a data
        self.parent.table.deleteRows(diff, startPos)
        self.parent.table.lastFilledRow -= diff

        # TODO Quando se deleta a ultima data, rowLength ja esta 0 e não é mostrada a mensagem de remoção com sucesso.
        if self.parent.table.rowLength <= 0:
            self.message.SetLabel('A tabela não contém nenhuma data.')
            self.yesBtn.Enable(False)
            self.parent.lastDate = ''
            msg = wx.MessageDialog(self, f'Dia {previousDate} removido da tabela com sucesso.', 'Sucesso', wx.OK | wx.ICON_INFORMATION)
            msg.ShowModal()

        else:
            self.parent.lastDate = self.parent.table.grid.GetCellValue(self.parent.table.lastFilledRow - 1, 0)
            self.message.SetLabel(f'A última data inserida é de {self.parent.lastDate}.\nVocê deseja removê-la?')
            msg = wx.MessageDialog(self, f'Dia {previousDate} removido da tabela com sucesso.', 'Sucesso', wx.OK | wx.ICON_INFORMATION)
            msg.ShowModal()


        self.parent.isSaved = False
        self.parent.updateTitleName()


    def OnCancel(self, event):
        ''' Chamada quando o botao 'Cancelar' for clicado. '''

        self.OnCloseWindow(None)


    def OnCloseWindow(self, event):
        ''' Funcao chamada quando o usuário clica no botao de fechar no canto superior direito. '''

        self.Destroy()
        self.parent.dataWindow = None